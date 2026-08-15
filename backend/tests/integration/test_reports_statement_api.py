import io
import uuid
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
import pytest
from httpx import AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.report_export.registry import registry as export_registry
from app.modules.auth.constants import AccountStatus
from app.modules.auth.models import User
from app.modules.auth.security import create_access_token, hash_password
from app.modules.companies.models import Company
from app.modules.invoices.constants import InvoiceStatus
from app.modules.invoices.models import Invoice
from app.modules.purchase.constants import PurchaseStatus
from app.modules.purchase.models import PurchaseBill
from app.modules.suppliers.models import Supplier

SUPER_ADMIN_EMAIL = "admin@fisherp.local"
SUPER_ADMIN_PASSWORD = "Admin@123"


async def _admin_headers(client: AsyncClient) -> dict[str, str]:
    response = await client.post(
        "/api/v1/auth/login",
        json={"email": SUPER_ADMIN_EMAIL, "password": SUPER_ADMIN_PASSWORD},
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


async def _admin_tenant_id(client: AsyncClient) -> uuid.UUID:
    response = await client.post(
        "/api/v1/auth/login",
        json={"email": SUPER_ADMIN_EMAIL, "password": SUPER_ADMIN_PASSWORD},
    )
    return uuid.UUID(response.json()["user"]["tenant_id"])


async def _make_user_headers(
    db_session: AsyncSession, tenant_id: uuid.UUID, permissions: list[str]
) -> dict[str, str]:
    user = User(
        tenant_id=tenant_id,
        email=f"user-{uuid.uuid4().hex[:8]}@fisherp.local",
        username=f"user-{uuid.uuid4().hex[:8]}",
        password_hash=hash_password("Whatever@123"),
        full_name="Test User",
        status=AccountStatus.ACTIVE,
        is_superuser=False,
    )
    db_session.add(user)
    await db_session.commit()
    token = create_access_token(
        subject=user.id, tenant_id=user.tenant_id, roles=["custom"], permissions=permissions
    )
    return {"Authorization": f"Bearer {token}"}


async def _make_company(
    db_session: AsyncSession, tenant_id: uuid.UUID, **overrides: Any
) -> Company:
    defaults: dict[str, Any] = {
        "tenant_id": tenant_id,
        "code": f"CO-{uuid.uuid4().hex[:8]}",
        "name": f"Company {uuid.uuid4().hex[:8]}",
        "company_type": "customer",
    }
    defaults.update(overrides)
    company = Company(**defaults)
    db_session.add(company)
    await db_session.commit()
    return company


async def _make_invoice(
    db_session: AsyncSession, tenant_id: uuid.UUID, company_id: uuid.UUID, **overrides: Any
) -> Invoice:
    defaults: dict[str, Any] = {
        "tenant_id": tenant_id,
        "company_id": company_id,
        "invoice_number": f"INV-{uuid.uuid4().hex[:8]}",
        "invoice_date": date(2026, 7, 1),
        "status": InvoiceStatus.ISSUED,
        "subtotal": Decimal("0"),
        "discount_amount": Decimal("0"),
        "taxable_amount": Decimal("0"),
        "tax_amount": Decimal("0"),
        "transport_charge": Decimal("0"),
        "other_charge": Decimal("0"),
        "round_off": Decimal("0"),
        "total_amount": Decimal("1000.00"),
        "paid_amount": Decimal("0"),
        "balance_amount": Decimal("1000.00"),
    }
    defaults.update(overrides)
    invoice = Invoice(**defaults)
    db_session.add(invoice)
    await db_session.commit()
    return invoice


async def _make_supplier(
    db_session: AsyncSession, tenant_id: uuid.UUID, **overrides: Any
) -> Supplier:
    defaults: dict[str, Any] = {
        "tenant_id": tenant_id,
        "code": f"SUP-{uuid.uuid4().hex[:8]}",
        "name": f"Supplier {uuid.uuid4().hex[:8]}",
    }
    defaults.update(overrides)
    supplier = Supplier(**defaults)
    db_session.add(supplier)
    await db_session.commit()
    return supplier


async def _make_purchase_bill(
    db_session: AsyncSession, tenant_id: uuid.UUID, supplier_id: uuid.UUID, **overrides: Any
) -> PurchaseBill:
    defaults: dict[str, Any] = {
        "tenant_id": tenant_id,
        "supplier_id": supplier_id,
        "bill_number": f"PB-{uuid.uuid4().hex[:8]}",
        "bill_date": date(2026, 7, 1),
        "status": PurchaseStatus.POSTED,
        "subtotal": Decimal("0"),
        "discount_amount": Decimal("0"),
        "taxable_amount": Decimal("0"),
        "tax_amount": Decimal("0"),
        "transport_charge": Decimal("0"),
        "other_charge": Decimal("0"),
        "round_off": Decimal("0"),
        "total_amount": Decimal("1000.00"),
        "paid_amount": Decimal("0"),
        "balance_amount": Decimal("1000.00"),
    }
    defaults.update(overrides)
    bill = PurchaseBill(**defaults)
    db_session.add(bill)
    await db_session.commit()
    return bill


class TestCustomerStatementAuth:
    async def test_requires_authentication(self, client: AsyncClient) -> None:
        response = await client.get("/api/v1/reports/customer-statement?format=excel")
        assert response.status_code == 401

    async def test_requires_reports_view_permission(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        headers = await _make_user_headers(db_session, tenant_id, ["company:view"])

        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=excel",
            headers=headers,
        )
        assert response.status_code == 403
        assert response.json()["error"]["code"] == "AUTHORIZATION_ERROR"

    async def test_reports_view_permission_is_sufficient(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        headers = await _make_user_headers(db_session, tenant_id, ["reports:view"])

        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=excel",
            headers=headers,
        )
        assert response.status_code == 200


class TestCustomerStatementValidation:
    async def test_missing_customer_id_is_422(self, client: AsyncClient) -> None:
        headers = await _admin_headers(client)
        response = await client.get(
            "/api/v1/reports/customer-statement?format=excel", headers=headers
        )
        assert response.status_code == 422

    async def test_unknown_customer_id_is_404(self, client: AsyncClient) -> None:
        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={uuid.uuid4()}&format=excel",
            headers=headers,
        )
        assert response.status_code == 404

    async def test_csv_format_is_rejected_with_clean_validation_error(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        headers = await _admin_headers(client)

        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=csv",
            headers=headers,
        )
        assert response.status_code == 422
        assert response.json()["error"]["code"] == "UNSUPPORTED_EXPORT_FORMAT"

    async def test_unknown_format_is_422(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        headers = await _admin_headers(client)

        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=xyz",
            headers=headers,
        )
        assert response.status_code == 422
        assert response.json()["error"]["code"] == "UNSUPPORTED_EXPORT_FORMAT"


class TestCustomerStatementExcel:
    async def test_excel_reflects_opening_and_closing_balance_and_preserves_running_balance(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id, name="Statement Excel Co")
        await _make_invoice(
            db_session,
            tenant_id,
            company.id,
            invoice_date=date(2026, 7, 10),
            total_amount=Decimal("4200.00"),
            balance_amount=Decimal("4200.00"),
        )

        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=excel",
            headers=headers,
        )

        assert response.status_code == 200
        assert "spreadsheetml" in response.headers["content-type"]
        assert ".xlsx" in response.headers["content-disposition"]

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        values = [
            cell.value for row in worksheet.iter_rows() for cell in row if cell.value is not None
        ]
        assert "Customer Statement" in values
        assert "Statement Excel Co" in " ".join(str(v) for v in values)
        assert Decimal("4200.00") in values
        assert "This is a system generated statement." in values

    async def test_excel_columns_do_not_include_transaction_type(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        await _make_invoice(db_session, tenant_id, company.id)

        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=excel",
            headers=headers,
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        values = [
            cell.value for row in worksheet.iter_rows() for cell in row if cell.value is not None
        ]
        assert "Reference Number" in values
        assert "Transaction Type" not in values


class TestCustomerStatementPDF:
    async def test_pdf_download(self, client: AsyncClient, db_session: AsyncSession) -> None:
        if not export_registry.is_registered("pdf"):
            pytest.skip("WeasyPrint native libraries are unavailable in this environment")

        tenant_id = await _admin_tenant_id(client)
        company = await _make_company(db_session, tenant_id)
        await _make_invoice(db_session, tenant_id, company.id)

        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/customer-statement?customer_id={company.id}&format=pdf",
            headers=headers,
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert response.content.startswith(b"%PDF")


class TestSupplierStatementAuth:
    async def test_requires_authentication(self, client: AsyncClient) -> None:
        response = await client.get("/api/v1/reports/supplier-statement?format=excel")
        assert response.status_code == 401

    async def test_requires_reports_view_permission(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        supplier = await _make_supplier(db_session, tenant_id)
        headers = await _make_user_headers(db_session, tenant_id, ["supplier:view"])

        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={supplier.id}&format=excel",
            headers=headers,
        )
        assert response.status_code == 403

    async def test_reports_view_permission_is_sufficient(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        supplier = await _make_supplier(db_session, tenant_id)
        headers = await _make_user_headers(db_session, tenant_id, ["reports:view"])

        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={supplier.id}&format=excel",
            headers=headers,
        )
        assert response.status_code == 200


class TestSupplierStatementValidation:
    async def test_missing_supplier_id_is_422(self, client: AsyncClient) -> None:
        headers = await _admin_headers(client)
        response = await client.get(
            "/api/v1/reports/supplier-statement?format=excel", headers=headers
        )
        assert response.status_code == 422

    async def test_unknown_supplier_id_is_404(self, client: AsyncClient) -> None:
        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={uuid.uuid4()}&format=excel",
            headers=headers,
        )
        assert response.status_code == 404

    async def test_csv_format_is_rejected_with_clean_validation_error(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        supplier = await _make_supplier(db_session, tenant_id)
        headers = await _admin_headers(client)

        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={supplier.id}&format=csv",
            headers=headers,
        )
        assert response.status_code == 422
        assert response.json()["error"]["code"] == "UNSUPPORTED_EXPORT_FORMAT"


class TestSupplierStatementExcel:
    async def test_excel_reflects_opening_and_closing_balance(
        self, client: AsyncClient, db_session: AsyncSession
    ) -> None:
        tenant_id = await _admin_tenant_id(client)
        supplier = await _make_supplier(db_session, tenant_id, name="Statement Excel Supplier")
        await _make_purchase_bill(
            db_session,
            tenant_id,
            supplier.id,
            bill_date=date(2026, 7, 5),
            total_amount=Decimal("3100.00"),
            balance_amount=Decimal("3100.00"),
        )

        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={supplier.id}&format=excel",
            headers=headers,
        )

        assert response.status_code == 200
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        values = [
            cell.value for row in worksheet.iter_rows() for cell in row if cell.value is not None
        ]
        assert "Supplier Statement" in values
        assert Decimal("3100.00") in values


class TestSupplierStatementPDF:
    async def test_pdf_download(self, client: AsyncClient, db_session: AsyncSession) -> None:
        if not export_registry.is_registered("pdf"):
            pytest.skip("WeasyPrint native libraries are unavailable in this environment")

        tenant_id = await _admin_tenant_id(client)
        supplier = await _make_supplier(db_session, tenant_id)
        await _make_purchase_bill(db_session, tenant_id, supplier.id)

        headers = await _admin_headers(client)
        response = await client.get(
            f"/api/v1/reports/supplier-statement?supplier_id={supplier.id}&format=pdf",
            headers=headers,
        )

        assert response.status_code == 200
        assert response.headers["content-type"] == "application/pdf"
        assert response.content.startswith(b"%PDF")
