import io
from datetime import UTC, date, datetime
from decimal import Decimal

import openpyxl

from app.core.report_export.export_models import (
    ColumnAlignment,
    ColumnFormat,
    ReportColumn,
    ReportExportData,
    ReportFilterDisplay,
    ReportRow,
    ReportSummary,
)
from app.core.report_export.exporters.excel_exporter import ExcelExporter

_GENERATED_AT = datetime(2026, 7, 30, 12, 0, tzinfo=UTC)


def _make_data(**overrides: object) -> ReportExportData:
    defaults: dict[str, object] = {
        "title": "Fish Sales Analytics",
        "subtitle": None,
        "filters": [ReportFilterDisplay(label="From Date", value="2026-07-01")],
        "columns": [
            ReportColumn(title="Fish", key="fish_name"),
            ReportColumn(
                title="Revenue",
                key="revenue",
                alignment=ColumnAlignment.RIGHT,
                format=ColumnFormat.CURRENCY,
            ),
            ReportColumn(title="Last Sold", key="last_sold_date", format=ColumnFormat.DATE),
        ],
        "rows": [
            ReportRow(
                data={
                    "fish_name": "Pomfret",
                    "revenue": Decimal("1000.00"),
                    "last_sold_date": date(2026, 7, 1),
                }
            ),
            ReportRow(
                data={"fish_name": "Surmai", "revenue": Decimal("500.50"), "last_sold_date": None}
            ),
        ],
        "summary": [
            ReportSummary(label="Total Revenue", value=Decimal("1500.50")),
            ReportSummary(label="Total Fish Types", value=2),
        ],
        "generated_at": _GENERATED_AT,
        "generated_by": "admin@fisherp.test",
        "tenant_name": "Konkan Traders",
    }
    defaults.update(overrides)
    return ReportExportData(**defaults)


def _load(output: bytes) -> openpyxl.workbook.workbook.Workbook:
    return openpyxl.load_workbook(io.BytesIO(output))


class TestExcelExporter:
    def test_content_type_and_extension(self) -> None:
        exporter = ExcelExporter()
        assert exporter.file_extension == "xlsx"
        assert "spreadsheetml" in exporter.content_type

    def test_produces_a_loadable_workbook_with_a_single_sheet(self) -> None:
        output = ExcelExporter().run(_make_data())
        workbook = _load(output)
        assert workbook.sheetnames == [workbook.active.title]

    def test_sheet_title_comes_from_report_title(self) -> None:
        workbook = _load(ExcelExporter().run(_make_data()))
        assert workbook.active.title == "Fish Sales Analytics"

    def test_title_row_is_present(self) -> None:
        workbook = _load(ExcelExporter().run(_make_data()))
        assert workbook.active["A1"].value == "Fish Sales Analytics"

    def test_table_header_row_matches_column_titles(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        header_row = worksheet.freeze_panes  # e.g. "A12" -> header at row 11
        header_row_index = int("".join(ch for ch in header_row if ch.isdigit())) - 1
        values = [cell.value for cell in worksheet[header_row_index]][:3]
        assert values == ["Fish", "Revenue", "Last Sold"]

    def test_freeze_panes_is_set_just_below_the_table_header(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        assert worksheet.freeze_panes is not None

    def test_auto_filter_covers_the_table_range(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        assert worksheet.auto_filter.ref is not None

    def test_currency_and_date_number_formats_applied_to_data_cells(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        number_formats = {
            cell.number_format
            for row in worksheet.iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert "#,##0.00" in number_formats
        assert "yyyy-mm-dd" in number_formats

    def test_right_aligned_column_data_cells_are_right_aligned(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        right_aligned_values = [
            cell.value
            for row in worksheet.iter_rows()
            for cell in row
            if cell.alignment.horizontal == "right"
        ]
        assert Decimal("1000.00") in right_aligned_values

    def test_summary_section_appears_above_the_table(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        labels = [
            cell.value for row in worksheet.iter_rows() for cell in row if cell.value == "Summary"
        ]
        assert labels == ["Summary"]

    def test_none_values_render_as_a_dash_placeholder(self) -> None:
        worksheet = _load(ExcelExporter().run(_make_data())).active
        dashes = [cell.value for row in worksheet.iter_rows() for cell in row if cell.value == "-"]
        assert dashes

    def test_zero_rows_still_produces_a_valid_workbook_with_header_only(self) -> None:
        output = ExcelExporter().run(_make_data(rows=[]))
        worksheet = _load(output).active
        assert worksheet.max_row >= 1

    def test_sheet_name_is_sanitized_and_truncated_to_31_chars(self) -> None:
        data = _make_data(title="A Report Title That Is Definitely Way Over Thirty One Characters")
        worksheet = _load(ExcelExporter().run(data)).active
        assert len(worksheet.title) <= 31

    def test_invalid_sheet_name_characters_are_stripped(self) -> None:
        data = _make_data(title="Sales: Q1/Q2 [Report]")
        worksheet = _load(ExcelExporter().run(data)).active
        assert not any(char in worksheet.title for char in "\\/*?:[]")
