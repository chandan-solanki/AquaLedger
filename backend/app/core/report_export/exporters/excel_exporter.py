import io
import re
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.report_export.base_exporter import BaseExporter
from app.core.report_export.export_models import ColumnAlignment, ColumnFormat, ReportExportData

_INVALID_SHEET_NAME_CHARS = re.compile(r"[\\/*?:\[\]]")
_MAX_COLUMN_WIDTH = 60
_MIN_COLUMN_WIDTH = 8

_NUMBER_FORMATS: dict[ColumnFormat, str] = {
    ColumnFormat.CURRENCY: "#,##0.00",
    ColumnFormat.DATE: "yyyy-mm-dd",
    ColumnFormat.DATETIME: "yyyy-mm-dd hh:mm",
    # A literal "%" suffix, not Excel's built-in 0.00% format - our
    # percent values already arrive pre-multiplied (e.g. Decimal("74.09")
    # meaning 74.09%), whereas 0.00% would re-multiply by 100.
    ColumnFormat.PERCENT: '0.00"%"',
    ColumnFormat.NUMBER: "#,##0.###",
}

_TITLE_FONT = Font(size=16, bold=True)
_SUBTITLE_FONT = Font(size=11, italic=True, color="595959")
_META_FONT = Font(size=9, color="808080")
_SECTION_HEADER_FONT = Font(bold=True)
_SUMMARY_LABEL_FONT = Font(bold=True)
_TABLE_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TABLE_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")


class ExcelExporter(BaseExporter):
    """Single-worksheet .xlsx export (TASKS.md Sprint 11 Session 5 Phase
    B): title/subtitle/generated-by preamble, an optional Filters
    section, an optional Summary section above the data table, then the
    table itself with bold white-on-blue headers, frozen just below the
    table's own header row, an auto filter over the table range, auto
    column widths, and per-column number formats (currency/date/percent/
    number) with numeric columns right-aligned. Cells keep their native
    Decimal/date/int type with an openpyxl `number_format` applied - never
    pre-rendered to text - so the workbook stays sortable/summable in
    Excel.
    """

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    file_extension = "xlsx"

    def export(self, data: ReportExportData) -> bytes:
        workbook = Workbook()
        worksheet = workbook.active
        assert worksheet is not None
        worksheet.title = self._sheet_name(data.title)

        row = self._write_preamble(worksheet, data)
        header_row = row
        self._write_table(worksheet, data, start_row=header_row)

        last_row = header_row + len(data.rows)
        last_col = len(data.columns)
        worksheet.freeze_panes = f"A{header_row + 1}"
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(last_col)}{max(last_row, header_row)}"
        )
        self._autofit_columns(worksheet, data, header_row=header_row)

        if data.footer:
            worksheet.cell(row=last_row + 2, column=1, value=data.footer).font = _META_FONT

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    @staticmethod
    def _sheet_name(title: str) -> str:
        sanitized = _INVALID_SHEET_NAME_CHARS.sub("", title)
        return sanitized[:31] or "Report"

    @classmethod
    def _write_preamble(cls, worksheet: Worksheet, data: ReportExportData) -> int:
        """Writes title/subtitle/generated-by/filters/summary above the
        table and returns the row number the table's own header row
        should start on."""
        row = 1

        worksheet.cell(row=row, column=1, value=data.title).font = _TITLE_FONT
        row += 1

        if data.subtitle:
            worksheet.cell(row=row, column=1, value=data.subtitle).font = _SUBTITLE_FONT
            row += 1

        generated_at = data.generated_at.strftime("%Y-%m-%d %H:%M")
        meta_cell = worksheet.cell(
            row=row, column=1, value=f"Generated {generated_at} by {data.generated_by}"
        )
        meta_cell.font = _META_FONT
        row += 1
        worksheet.cell(row=row, column=1, value=f"Tenant: {data.tenant_name}").font = _META_FONT
        row += 1

        if data.filters:
            row += 1
            worksheet.cell(row=row, column=1, value="Filters").font = _SECTION_HEADER_FONT
            row += 1
            for filter_display in data.filters:
                worksheet.cell(
                    row=row, column=1, value=filter_display.label
                ).font = _SUMMARY_LABEL_FONT
                worksheet.cell(row=row, column=2, value=filter_display.value)
                row += 1

        if data.summary:
            row += 1
            worksheet.cell(row=row, column=1, value="Summary").font = _SECTION_HEADER_FONT
            row += 1
            for summary_item in data.summary:
                worksheet.cell(
                    row=row, column=1, value=summary_item.label
                ).font = _SUMMARY_LABEL_FONT
                value = summary_item.value
                if isinstance(value, Decimal):
                    cell = worksheet.cell(row=row, column=2, value=value)
                    cell.number_format = "#,##0.00"
                elif isinstance(value, int):
                    cell = worksheet.cell(row=row, column=2, value=value)
                    cell.number_format = "#,##0"
                else:
                    worksheet.cell(
                        row=row, column=2, value=str(value) if value is not None else "-"
                    )
                row += 1

        row += 1
        return row

    @staticmethod
    def _write_table(worksheet: Worksheet, data: ReportExportData, *, start_row: int) -> None:
        for col_index, column in enumerate(data.columns, start=1):
            cell = worksheet.cell(row=start_row, column=col_index, value=column.title)
            cell.font = _TABLE_HEADER_FONT
            cell.fill = _TABLE_HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_offset, row in enumerate(data.rows, start=1):
            row_index = start_row + row_offset
            for col_index, column in enumerate(data.columns, start=1):
                value = row.data[column.key]
                cell = worksheet.cell(
                    row=row_index, column=col_index, value=value if value is not None else "-"
                )
                if value is not None:
                    number_format = _NUMBER_FORMATS.get(column.format)
                    if number_format:
                        cell.number_format = number_format
                    if column.alignment == ColumnAlignment.RIGHT:
                        cell.alignment = Alignment(horizontal="right")
                    elif column.alignment == ColumnAlignment.CENTER:
                        cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _autofit_columns(worksheet: Worksheet, data: ReportExportData, *, header_row: int) -> None:
        for col_index, column in enumerate(data.columns, start=1):
            longest = len(column.title)
            for row in data.rows:
                value = row.data[column.key]
                longest = max(longest, len(str(value)) if value is not None else 1)
            width = max(_MIN_COLUMN_WIDTH, min(_MAX_COLUMN_WIDTH, longest + 2))
            worksheet.column_dimensions[get_column_letter(col_index)].width = width
